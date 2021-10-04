import openpyxl 
path = "roaddistance.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row
column = sheet_obj.max_column
print("Total Rows:", row)
print("Total Columns:", column)
for i in range(1, column): 
    cell_obj = sheet_obj.cell(row = 2, column = i) 
    # print(cell_obj.value, end = " ")
    for j in range(1, row + 1): 
        cell_obj2 = sheet_obj.cell(row = j, column = 2) 
        print("data(", end = "" )
        print(cell_obj.value, end = ", ")
        print(cell_obj2.value, end = ", ")
        cell_obj3 = sheet_obj.cell(row = j, column = i) 
        print(cell_obj3.value, end = ") ") 
        print("")